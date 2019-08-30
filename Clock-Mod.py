import PySimpleGUI as sg
from sys import exit
from pandas import DataFrame
from pandas import ExcelWriter
import datetime
import openpyxl
from openpyxl import load_workbook
import time


writer = ExcelWriter(r'test.xlsx', engine='openpyxl')
book = load_workbook(r'test.xlsx')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}
check1 = False
check2 = False
check3 = False
inTime1 = 0
outTime1 = 0
inTime2 = 0
outTime2 = 0
inTime3 = 0
outTime3 = 0
str=''
str2=''
layout = [[sg.Image(filename='logo.png', pad=(120,20))],
          [sg.Text(str, background_color='white',pad=(230,1), key='codeinput', font=("Helvetica", 25), size=(4,1))],
          [sg.RButton('1',button_color=('white', '#2d2d2d'),size=(20,7)), sg.RButton('2',button_color=('white', '#2d2d2d'),size=(20,7)), sg.RButton('3',button_color=('white', '#2d2d2d'),size=(20,7))],
          [sg.RButton('4',button_color=('white', '#2d2d2d'),size=(20,7)), sg.RButton('5',button_color=('white', '#2d2d2d'),size=(20,7)), sg.RButton('6',button_color=('white', '#2d2d2d'),size=(20,7))],
          [sg.RButton('7',button_color=('white', '#2d2d2d'),size=(20,7)), sg.RButton('8',button_color=('white', '#2d2d2d'),size=(20,7)), sg.RButton('9',button_color=('white', '#2d2d2d'),size=(20,7))],
          [sg.RButton('Submit',button_color=('white', '#15438e'),size=(20,7)), sg.RButton('Clear',button_color=('white', '#2d2d2d'),size=(20,7)), sg.RButton('Exit',button_color=('white', '#15438e'),size=(20,7))],
          [sg.Text(str2, background_color='white',pad=(90,1), key='codeoutput', font=("Helvetica", 25), size=(130,1))],]
window = sg.Window('Employee Clock', auto_size_text=True, default_element_size=(100, 1), background_color='white', resizable=True, icon='logo.ico', no_titlebar=True, grab_anywhere=True, size=(580,950)).Layout(layout)

while True:
    event, values = window.Read()

    if event is None or event == 'Exit':
        break
    if (event=='1' or event=='2' or event=='3' or event=='4' or event=='5' or event=='6' or event=='7' or event=='8' or event=='9') and len(str)<4:
        str = str + event
        window.Element('codeinput').Update(value=str)

    if event=='Clear':
        str=''
        window.Element('codeinput').Update(value=str)
        str2=''
        window.Element('codeoutput').Update(value=str)
    if event=='Submit':
        text = str
        if text == "1111":
            if check1 == False:
                hoursclock = int(datetime.datetime.now().strftime("%H"))
                minsclock = int(datetime.datetime.now().strftime("%M"))/60
                inTime1 = hoursclock + minsclock
                df = DataFrame({'Data': [datetime.datetime.now().strftime("%Y-%m-%d")],
                                'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "IN"})
                df.to_excel(writer, sheet_name='Dispatch', index=False, startrow=writer.sheets['Dispatch'].max_row,
                            header=False)
                check1 = True
                writer.save()
                str2 = "Clocked In at " + datetime.datetime.now().strftime("%I:%M %p")
                window.Element('codeoutput').Update(value=str2)
                str = ''
                window.Element('codeinput').Update(value=str)

            elif check1 == True:
                hoursclock = int(datetime.datetime.now().strftime("%H"))
                minsclock = int(datetime.datetime.now().strftime("%M"))/60
                outTime1 = hoursclock + minsclock
                timepayable = outTime1 - inTime1
                inTime1 = 0
                outTime1 = 0
                df = DataFrame({'Data': [""], 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "OUT",
                                'Data3': timepayable})
                df.to_excel(writer, sheet_name='Dispatch', index=False, startrow=writer.sheets['Dispatch'].max_row,
                            header=False)
                check1 = False
                writer.save()
                str2 = "Clocked Out at " + datetime.datetime.now().strftime("%I:%M %p")
                window.Element('codeoutput').Update(value=str2)
                str = ''
                window.Element('codeinput').Update(value=str)

        if text == "1337":
            if check2 == False:
                hoursclock = int(datetime.datetime.now().strftime("%H"))
                minsclock = int(datetime.datetime.now().strftime("%M"))/60
                inTime2 = hoursclock + minsclock
                df1 = DataFrame({'Data': [datetime.datetime.now().strftime("%Y-%m-%d")],
                                 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "IN"})
                df1.to_excel(writer, sheet_name='Inder', index=False, startrow=writer.sheets['Inder'].max_row,
                             header=False)
                check2 = True
                writer.save()
                str2 = "Clocked In at " + datetime.datetime.now().strftime("%I:%M %p")
                window.Element('codeoutput').Update(value=str2)

                str = ''
                window.Element('codeinput').Update(value=str)
            elif check2 == True:
                hoursclock = int(datetime.datetime.now().strftime("%H"))
                minsclock = int(datetime.datetime.now().strftime("%M"))/60
                outTime2 = hoursclock + minsclock
                timepayable = outTime2 - inTime2
                inTime2 = 0
                outTime2 = 0
                df1 = DataFrame({'Data': [""], 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "OUT",
                                 'Data3': timepayable})
                df1.to_excel(writer, sheet_name='Inder', index=False, startrow=writer.sheets['Inder'].max_row,
                             header=False)
                check2 = False
                writer.save()
                str2 = "Clocked Out at " + datetime.datetime.now().strftime("%I:%M %p")
                window.Element('codeoutput').Update(value=str2)
                str = ''
                window.Element('codeinput').Update(value=str)

        if text == "2222":
            if check3 == False:
                hoursclock = int(datetime.datetime.now().strftime("%H"))
                minsclock = int(datetime.datetime.now().strftime("%M"))/60
                inTime3 = hoursclock + minsclock
                df2 = DataFrame({'Data': [datetime.datetime.now().strftime("%Y-%m-%d")],
                                 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "IN"})
                df2.to_excel(writer, sheet_name='Safety', index=False, startrow=writer.sheets['Safety'].max_row,
                             header=False)
                check3 = True
                writer.save()
                str2 = "Clocked In at " + datetime.datetime.now().strftime("%I:%M %p")
                window.Element('codeoutput').Update(value=str2)
                str = ''
                window.Element('codeinput').Update(value=str)
            elif check3 == True:
                hoursclock = int(datetime.datetime.now().strftime("%H"))
                minsclock = int(datetime.datetime.now().strftime("%M"))/60
                outTime3 = hoursclock + minsclock
                timepayable = outTime3 - inTime3
                inTime3 = 0
                outTime3 = 0
                df2 = DataFrame({'Data': [datetime.datetime.now().strftime("%Y-%m-%d")],
                                 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "OUT",
                                 'Data3': timepayable})
                df2.to_excel(writer, sheet_name='Safety', index=False, startrow=writer.sheets['Safety'].max_row,
                             header=False)
                check3 = False
                writer.save()
                str2 = "Clocked Out at " + datetime.datetime.now().strftime("%I:%M %p")
                window.Element('codeoutput').Update(value=str2)
                str = ''
                window.Element('codeinput').Update(value=str)




window.Close()

