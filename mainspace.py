from tkinter import *
from pandas import DataFrame
from pandas import ExcelWriter
import datetime
import openpyxl
from openpyxl import load_workbook
root = Tk()
root.title("Travel Transport Employee Clock")
root.geometry("600x600")
writer = ExcelWriter(r'//mycloudex2ultra/Public/EmployeeClock/Clock.xlsx', engine='openpyxl')
book = load_workbook(r'//mycloudex2ultra/Public/EmployeeClock/Clock.xlsx')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}
check1 = False
check2 = False
check3 = False
inTime1 = 0
outTime1 = 0
inTime2 = 0
outTime2 = 0
inTime3 = 0
outTime3 = 0
#Functions
def enterval():
    global check1
    global check2
    global check3
    global inTime1
    global outTime1
    global inTime2
    global outTime2
    global inTime3
    global outTime3
    text = empcode.get()

    if text == "1111":
        if check1 == False:
            hoursclock = int(datetime.datetime.now().strftime("%H"))
            minsclock = int(datetime.datetime.now().strftime("%M"))
            if minsclock <= 15 and minsclock > 0:
                minsclock = 15/60
            elif minsclock > 15 and minsclock <= 30:
                minsclock = 30/60
            elif minsclock > 30 and minsclock <= 45:
                minsclock = 45/60
            elif minsclock > 45 or minsclock == 0:
                minsclock = 0
                hoursclock = hoursclock+1
            inTime1 = hoursclock + minsclock
            df = DataFrame({'Data': [datetime.datetime.now().strftime("%Y-%m-%d")], 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "IN"})
            df.to_excel(writer, sheet_name='Dispatch', index=False, startrow=writer.sheets['Dispatch'].max_row, header=False)
            check1 = True
            label1["text"]= "Clocked In at " + datetime.datetime.now().strftime("%I:%M %p")
            writer.save()
            empcode.delete(0,END)
            return
        if check1 == True:
            hoursclock = int(datetime.datetime.now().strftime("%H"))
            minsclock = int(datetime.datetime.now().strftime("%M"))
            if minsclock < 15:
                minsclock = 0
            elif minsclock >= 15 and minsclock < 30:
                minsclock = 15/60
            elif minsclock >=30 and minsclock < 45:
                minsclock = 30/60
            elif minsclock >=45:
                minsclock = 45/60
            outTime1 = hoursclock + minsclock
            timepayable = outTime1 - inTime1
            inTime1 = 0
            outTime1 = 0
            df = DataFrame({'Data': [""], 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "OUT", 'Data3': timepayable})
            df.to_excel(writer, sheet_name='Dispatch', index=False, startrow=writer.sheets['Dispatch'].max_row, header=False)
            check1 = False
            label1["text"] = "Clocked Out at " + datetime.datetime.now().strftime("%I:%M %p")
            writer.save()
            empcode.delete(0, END)
            return

    if text == "6704":
            if check2 ==False:
                hoursclock = int(datetime.datetime.now().strftime("%H"))
                minsclock = int(datetime.datetime.now().strftime("%M"))
                if minsclock <= 15 and minsclock > 0:
                    minsclock = 15 / 60
                elif minsclock > 15 and minsclock <= 30:
                    minsclock = 30 / 60
                elif minsclock > 30 and minsclock <= 45:
                    minsclock = 45 / 60
                elif minsclock > 45 or minsclock == 0:
                    minsclock = 0
                    hoursclock = hoursclock + 1
                inTime2 = hoursclock + minsclock
                df1 = DataFrame({'Data': [datetime.datetime.now().strftime("%Y-%m-%d")], 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "IN"})
                df1.to_excel(writer, sheet_name='Inder', index=False, startrow=writer.sheets['Inder'].max_row, header=False)
                check2 = True
                label1["text"] = "Clocked In at " + datetime.datetime.now().strftime("%I:%M %p")
                writer.save()
                empcode.delete(0,END)
                return
            if check2 ==True:
                hoursclock = int(datetime.datetime.now().strftime("%H"))
                minsclock = int(datetime.datetime.now().strftime("%M"))
                if minsclock < 15:
                    minsclock = 0
                elif minsclock >= 15 and minsclock < 30:
                    minsclock = 15 / 60
                elif minsclock >= 30 and minsclock < 45:
                    minsclock = 30 / 60
                elif minsclock >= 45:
                    minsclock = 45 / 60
                outTime2 = hoursclock + minsclock
                timepayable = outTime2 - inTime2
                inTime2 = 0
                outTime2 = 0
                df1 = DataFrame({'Data': [""], 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "OUT", 'Data3': timepayable})
                df1.to_excel(writer, sheet_name='Inder', index=False, startrow=writer.sheets['Inder'].max_row,header=False)
                check2 = False
                label1["text"] = "Clocked Out at " + datetime.datetime.now().strftime("%I:%M %p")
                writer.save()
                empcode.delete(0, END)
                return

    if text == "2222":
            if check3 == False:
                hoursclock = int(datetime.datetime.now().strftime("%H"))
                minsclock = int(datetime.datetime.now().strftime("%M"))
                if minsclock <= 15 and minsclock > 0:
                    minsclock = 15 / 60
                elif minsclock > 15 and minsclock <= 30:
                    minsclock = 30 / 60
                elif minsclock > 30 and minsclock <= 45:
                    minsclock = 45 / 60
                elif minsclock > 45 or minsclock == 0:
                    minsclock = 0
                    hoursclock = hoursclock + 1
                inTime3 = hoursclock + minsclock
                df2 = DataFrame({'Data': [datetime.datetime.now().strftime("%Y-%m-%d")], 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "IN"})
                df2.to_excel(writer, sheet_name='Safety', index=False, startrow=writer.sheets['Safety'].max_row,header=False)
                check3=True
                label1["text"] = "Clocked In at " + datetime.datetime.now().strftime("%I:%M %p")
                writer.save()
                empcode.delete(0, END)
                return
            if check3 ==True:
                hoursclock = int(datetime.datetime.now().strftime("%H"))
                minsclock = int(datetime.datetime.now().strftime("%M"))
                if minsclock < 15:
                    minsclock = 0
                elif minsclock >= 15 and minsclock < 30:
                    minsclock = 15 / 60
                elif minsclock >= 30 and minsclock < 45:
                    minsclock = 30 / 60
                elif minsclock >= 45:
                    minsclock = 45 / 60
                outTime3 = hoursclock + minsclock
                timepayable = outTime3 - inTime3
                inTime3 = 0
                outTime3 = 0
                df2 = DataFrame({'Data': [datetime.datetime.now().strftime("%Y-%m-%d")], 'Data1': [datetime.datetime.now().strftime("%H:%M")], 'Data2': "OUT", 'Data3': timepayable})
                df2.to_excel(writer, sheet_name='Safety', index=False, startrow=writer.sheets['Safety'].max_row, header=False)
                check3 = False
                label1["text"] = "Clocked Out at " + datetime.datetime.now().strftime("%I:%M %p")
                writer.save()
                empcode.delete(0, END)
                return
    else:
        return

def del_text():
    empcode.delete(0,END)


def set_text(text):
    empcode.insert(END,text)
    return

#GUI
root["bg"] = "white"
Label(root, text="Code:", font="Roboto", bg="white").grid(row=1, column=2, sticky=NSEW)
empcode = Entry(root, justify=CENTER, bg="#dcdcdc", width=80, borderwidth=0)
empcode.grid(row=2,column=1, sticky=NSEW, columnspan=3,pady=10)
#Good way might work on it later
#numbers = [['1', '2', '3'],\
#        ['4', '5', '6'],\
#         ['7', '8', '9']]
#for i, (x, y, z) in enumerate(numbers):
#    Button(root, text=x, command=lambda:set_text(x)).grid(row=i+3, column=1, ipadx=60, ipady=30, sticky=NSEW)
#    Button(root, text=y, command=lambda:set_text(y)).grid(row=i+3, column=2, ipadx=60, ipady=30,sticky=NSEW)
#    Button(root, text=z, command=lambda:set_text(z)).grid(row=i+3, column=3, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="1", command=lambda:set_text("1")).grid(row=3, column=1, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="2", command=lambda:set_text("2")).grid(row=3, column=2, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="3", command=lambda:set_text("3")).grid(row=3, column=3, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="4", command=lambda:set_text("4")).grid(row=4, column=1, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="5", command=lambda:set_text("5")).grid(row=4, column=2, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="6", command=lambda:set_text("6")).grid(row=4, column=3, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="7", command=lambda:set_text("7")).grid(row=5, column=1, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="8", command=lambda:set_text("8")).grid(row=5, column=2, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="9", command=lambda:set_text("9")).grid(row=5, column=3, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="0", command=lambda:set_text("0")).grid(row=6, column=2, ipadx=60, ipady=30, sticky=NSEW)
Button(root, text="Submit", command=enterval).grid(row=6, column=3, ipadx=60, ipady=30,sticky=NSEW)
Button(root, text="Clear", command=del_text).grid(row=6, column=1, ipadx=60, ipady=30,sticky=NSEW)
label1 = Label(root, text="", font=("Roboto", 38), bg="white")
label1.grid(row=7,column=1,sticky=NSEW, columnspan=3)

root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(7, weight=1)
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(4, weight=1)
root.mainloop()